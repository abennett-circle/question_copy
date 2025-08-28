# Last Updated: 2025-08-28
# Description: A class for managing reference and unanswered questionnaires for Compliance team.
#              (1) Creates Reference_Questionnaire and Unanswered_Questionnaire objects.
#              (2) Matches questions to the reference questionnaire.

################################################################################
# Imports
################################################################################
from .Reference_Questionnaire import Reference_Questionnaire
from .Unanswered_Questionnaire import Unanswered_Questionnaire
from openai import OpenAI
import httpx
import json
import os
from dotenv import load_dotenv
import pandas as pd
from openpyxl.styles import PatternFill, Alignment

################################################################################
# Questionnaire Filler Class
################################################################################

"""
Purpose: Creates and manages reference and unanswered questionnaire objects.

Attributes:
    reference_questionnaire (Reference_Questionnaire): The reference questionnaire object.
    unanswered_questionnaire (Unanswered_Questionnaire): The unanswered questionnaire object.
    ai_client: The OpenAI client.

Methods:
    N/A
"""
class Questionnaire_Filler(object):

    """Initialize the Questionnaire_Filler class.
        
        Args:
            reference_file_name (str): Path to the reference questionnaire file.
            reference_question_col (str): Column name for questions in reference file.
            reference_answer_col (str): Column name for answers in reference file.
            unanswered_file_name (str): Path to the unanswered questionnaire file.
            unanswered_question_col (str): Column name for questions in unanswered file.
            unanswered_answer_col (str): Column name for answers in unanswered file.
            ai_url (str, optional): AI API base URL. If not provided, uses CHATAI_BASE_URL from config.env.
            api_key (str, optional): AI API key. If not provided, uses CHATAI_API_KEY from config.env.
            
        Environment Variables (config.env):
            CHATAI_BASE_URL: ChatAI Circle API base URL
            CHATAI_API_KEY: ChatAI Circle API key
    """
    def __init__(self, reference_file_name,
                       reference_question_col,
                       reference_answer_col,
                       unanswered_file_name,
                       unanswered_question_col,
                       unanswered_answer_col,
                       default_model="gpt-4o",
                       ai_url=None,
                       api_key=None,
                       accuracy_threshold=0.85):
        
        # Load environment variables - try multiple paths
        config_paths = ['config.env', '../config.env', './config.env']
        for config_path in config_paths:
            if os.path.exists(config_path):
                load_dotenv(config_path)
                break
        
        # Set AI credentials from environment or parameters
        self.ai_url = ai_url or os.getenv('CHATAI_BASE_URL')
        self.api_key = api_key or os.getenv('CHATAI_API_KEY')
        self.default_model = default_model    
        # Creates questionnaire objects
        self.reference_questionnaire = Reference_Questionnaire(
            file_path=reference_file_name,
            question_col=reference_question_col,
            answer_col=reference_answer_col
        )
        
        # Creates the unanswered questionnaire object
        self.unanswered_questionnaire = Unanswered_Questionnaire(
            file_path=unanswered_file_name,
            question_col=unanswered_question_col,
            answer_col=unanswered_answer_col
        )

        # Builds the AI client
        if self.ai_url and self.api_key:
            self.ai_client = self._build_ai_client(self.ai_url, self.api_key)
        else:
            raise ValueError("AI URL and API key must be provided either as parameters or in config.env file")

        # Set the accuracy threshold
        self.accuracy_threshold = accuracy_threshold

        # Determines the static compliance match, hardcoded for requested questions by compliance team
        self.static_compliance_matches = {"What is the most sensitive data classification that the third party will have access to for this engagement?" : "Classification"}



    """Builds the AI client.
        
        Returns:
            client (OpenAI): The OpenAI client.
    """
    def _build_ai_client(self, ai_url, api_key):

        # Create a proper httpx client with increased timeout
        http_client = httpx.Client(
            timeout=60.0 
        )

        # Configure the OpenAI client with the proxy URL
        client = OpenAI(
            base_url=ai_url,
            api_key=api_key,
            http_client=http_client
        )

        # Returns the client
        return client
    


    """Makes an AI request.
        
        Args:
            user_prompt (str): The user prompt.
            model (str): The model to use.
            system_prompt (str): The system prompt.
            temperature (float): The temperature.
            max_tokens (int): The maximum number of tokens.
            has_arr_content (bool): Whether the request has array content.
            arr_content (dict): Message content for array content.
    """
    def _make_ai_request(self, 
                         user_prompt,
                         model="gpt-4o",
                         system_prompt="You are trapobot, a helpful assistant that fills out compliance questionnaires. ",
                         temperature=0.7,
                         max_tokens=500,
                         has_arr_content=False,
                         arr_content=dict()):
        
        # Stores the message content
        message_content = [{"role": "system", "content": system_prompt},
                           {"role": "user", "content": user_prompt}]
        
        # If the request has content, add it to the message content
        if (has_arr_content):
            message_content.append({"role": "user", "content": json.dumps(arr_content, ensure_ascii=False)})
        
        # Make a request
        response = self.ai_client.chat.completions.create(
            model=model,
            messages=message_content,
            temperature=temperature,
            max_tokens=max_tokens,
            response_format={"type": "json_object"}
        )

        # Returns the response
        return response



    """Matches exact questions to the reference questionnaire and resets their reference question ID.
        
        Returns:
            unmatched_questions_remaining (list): The remaining unmatched questions.
            reference_questions_remaining (list): The remaining reference questions.
    """
    def _match_exact_questions(self):
        
        # Gets the questions from the reference questionnaire and the unanswered questionnaire
        reference_questions = self.reference_questionnaire.get_questions()
        unanswered_questions = self.unanswered_questionnaire.get_questions()

        # Creates sets of the dictionary keys to keep track of unmatched question pairs
        unanswered_questions_remaining = set(unanswered_questions.keys())
        reference_questions_remaining = set(reference_questions.keys())

        # Matches the question
        for question in unanswered_questions:

            # If question is found in the existing questionnaire
            if (question in reference_questions):

                # Gets the matching reference question
                matched_question = reference_questions[question].get_question()

                # Gets the answer from the reference questionnaire
                self.unanswered_questionnaire.questions[question].set_reference_question(matched_question)

                # Sets the question match score
                self.unanswered_questionnaire.questions[question].set_question_match_score(1)

                # Removes the question from the remaining questions
                unanswered_questions_remaining.discard(question)
                reference_questions_remaining.discard(matched_question)

            # If the question is in the static compliance match, set the reference question and question match score
            if (question in self.static_compliance_matches):

                # Sets the reference question and question match score
                self.unanswered_questionnaire.questions[question].set_reference_question(self.static_compliance_matches[question])
                self.unanswered_questionnaire.questions[question].set_question_match_score(1)

                # Removes the question from the remaining questions
                unanswered_questions_remaining.discard(question)
                reference_questions_remaining.discard(self.static_compliance_matches[question])

        # Returns the reamining unmatched questions and reference questions
        return (list(unanswered_questions_remaining), list(reference_questions_remaining))

    

    """Matches questions to the reference questionnaire.
        
        Returns:
            dict: Dictionary containing the AI-generated question matches.
    """
    def _match_questions_to_reference(self):

        # Static variables for the questionnaire matching task
        QUESTIONNAIRE_MATCHING_TEMPERATURE = 0
        QUESTIONNAIRE_MATCHING_MODEL = self.default_model

        # Stores the system prompt for the questionnaire matching task
        QUESTIONNAIRE_MATCHING_PROMPT = ("You are QuestionnaireMatcher. Given a new question + answered candidates, return MATCH if exactly same intent, "
                                              "else NO_MATCH. SAME intent = same subject + attribute + scope (timeframe, jurisdiction, quantity, polarity). "
                                              "Be conservative: prefer NO_MATCH if unsure. Do not match general vs. specific (dog≠pets, car≠vehicle), different "
                                              "timeframe/jurisdiction/thresholds, or opposite polarity. Paraphrases ok if identical meaning. Input JSON: "
                                              "{new_question, candidates, hints?}. Output JSON: {\"decision\":\"MATCH|NO_MATCH\",\"match_id\":\"<id|null>\","
                                              "\"confidence\":0-100,\"rationale\":\"...\",\"key_entities\":{\"subject\":\"...\",\"attribute\":\"...\",\"scope\":"
                                              "{\"timeframe\":\"...\",\"jurisdiction\":\"...\",\"quantity_or_units\":\"...\",\"polarity\":\"positive|negative|neutral\"}}}. "
                                              "Only MATCH if confidence≥70. Steps: 1) Normalize/extract. 2) If hard conflict, NO_MATCH. 3) Else compute similarity; "
                                              "require explicit subject match. 4) If similarity/conf<49, NO_MATCH else MATCH. Examples MATCH: \"Do you have a cat?\"↔"
                                              "\"Check yes if you have a cat.\", \"Currently employed full-time?\"↔\"Have a full-time job right now?\". Examples NO_MATCH: "
                                              "\"Dog?\"↔\"Pets?\", \"Currently insured?\"↔\"Ever had insurance?\", \"Registered in CA?\"↔\"Registered?\"." )

        # Stores the instructions for the questionnaire matching task
        USER_PROMPT = """You will receive:
                        1) A list of UNMATCHED questions for the unanswered questionnaire.
                        2) A list of REFERENCE_CATALOG of questions for the reference questionnaire.

                        TASK:
                        For each UNMATCHED question, select the best-matching question from REFERENCE_CATALOG.

                        Rules:
                        - CRITICAL: Return ONLY valid JSON with this exact structure. Ensure all strings are properly escaped:
                        {
                            "matches": [
                                {
                                    "unmatched_question": "string",
                                    "match": {
                                        "matched_question": "string",
                                        "similarity": 0.85,
                                        "justification": "string"
                                    },
                                    "no_match": false
                                }
                            ]
                        }

                        Constraints:
                        - Keep justification one sentence, no hedging.
                        - If no adequate match exists, set no_match=true and match=None.
                        - Be deterministic.
                        - similarity should be 0.00 to 1.00
                        - Must return valid JSON object with "matches" array

                        Evaluation hints: Decompose into subject/attribute/scope; require explicit subject match; reject hypernym/hyponym jumps; align timeframe/jurisdiction/thresholds/units; enforce same polarity; don't match general↔specific; require full alignment for multi-part questions; normalize acronyms only if supported by hints; distinguish policy vs. practice and state vs. proof; treat data types distinctly (PII/PHI/telemetry). Hard NOs: opposite polarity, numeric or timeframe conflict, jurisdiction mismatch, storage vs. transport, collect vs. retain vs. delete, DPIA vs. control. Confidence starts at semantic similarity minus penalties; MATCH only if ≥49. For ties, prefer exact scope and threshold matches, then fewer assumptions."""
        
        # Stores the unmatched questions
        unanswered_questions_remaining, reference_questions_remaining = self._match_exact_questions()

        # Stores the question payload for the AI request
        question_payload = {
            "unanswered_questions": unanswered_questions_remaining,
            "reference_questions": reference_questions_remaining
        }

        # Makes an AI request with the unmatched questions and reference questions
        response = self._make_ai_request(
            user_prompt=USER_PROMPT,
            model=QUESTIONNAIRE_MATCHING_MODEL,
            system_prompt=QUESTIONNAIRE_MATCHING_PROMPT,
            temperature=QUESTIONNAIRE_MATCHING_TEMPERATURE,
            max_tokens=2000,
            has_arr_content=True,
            arr_content=question_payload
        )


        # Gets the response content
        resp_content = response.choices[0].message.content
        
        # Parse response content using OpenAI's native JSON parsing
        try:
            # Parse the JSON response directly (OpenAI's JSON mode ensures valid JSON)
            response_data = json.loads(resp_content)
            # Extract the matches array from the structured response
            matches_array = response_data.get("matches", [])
            print(f"Successfully parsed {len(matches_array)} question matches")

            return matches_array
        
        # Errors if the JSON returned by the AI cleint is not valid
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response from AI model: {e}")
            print(f"Raw response content: {resp_content[:200]}...")
            return []
        except Exception as e:
            print(f"Unexpected error during response processing: {e}")
            print(f"Error type: {type(e).__name__}")
            print(f"Error message: {e}")
            return []
    


    """Helper method to check if text contains meaningful content (letters or numbers).
        
        Args:
            text (any): The text/value to check.
            
        Returns:
            bool: True if text contains alphanumeric characters, False otherwise.
    """
    def _has_meaningful_content(self, text):
        if (text is None):
            return False
        
        # Check for NaN values (pandas NaN becomes nan when converted to string)
        if (pd.isna(text)):
            return False
        
        # Convert everything to string first
        text_str = str(text)
        
        # Additional check for string representations of NaN/null values
        if (text_str.lower() in ['nan', 'none', 'null', '']):
            return False
        
        # Strip whitespace and check if there are any alphanumeric characters
        stripped = text_str.strip()

        # Returns True if the text contains alphanumeric characters, False otherwise
        return len(stripped) > 0 and any(c.isalnum() for c in stripped)


    """ Gets the items for the answer matching task.
        
        Returns:
            items_json (list): The items for the answer matching task.
    """
    def _get_items_for_answer_matching(self):

        # Gets the questions from the reference questionnaire and the unanswered questionnaire
        items_json = []

        # Goes through the unanswered questionnaire
        for question in self.unanswered_questionnaire.get_questions():

            # Gets the current answer
            current_answer = self.unanswered_questionnaire.questions[question].get_answer()

            # Gets the reference question
            reference_question = self.unanswered_questionnaire.questions[question].get_reference_question()

            # Only process if both reference question and current answer have meaningful content
            if (reference_question != None and reference_question != "" and self._has_meaningful_content(current_answer)):

                # Gets the reference answer
                reference_answer = self.reference_questionnaire.questions[reference_question].get_answer()

                # If the answers are a match, set the answer match score to 1
                if (str(current_answer).strip() == str(reference_answer).strip()):
                    self.unanswered_questionnaire.questions[question].set_answer_match_score(1)

                # Otherwise, add the item to the items_json list
                else:
                    items_json.append({
                        "q": question,
                        "a1": current_answer,
                        "a2": reference_answer
                    })

        # Returns the items_json list
        return items_json
    
    

    """Fills the answer match scores for each set of answers.
        
        Returns:
            answer_scores (dict): Dictionary mapping questions to their answer match scores.
    """
    def _fill_answer_matches_score(self):

        # Gets the questions from the reference questionnaire and the unanswered questionnaire
        items_json = self._get_items_for_answer_matching()
        
        # If no items to process, return empty dict
        if not items_json:
            print("No items to process for answer matching")
            return {}
        
        print(f"Processing {len(items_json)} items for answer matching")

        # Static variables for the answer matching task
        ANSWER_MATCHING_TEMPERATURE = 0
        ANSWER_MATCHING_MODEL = self.default_model

        # Stores the system prompt for the answer matching task
        ANSWER_MATCHING_PROMPT = ("You are AnswerMatcher. Given two answers to a question, you will return the answer match score. "
                                  "The answer match score is a number between 0 and 1, where 1 is the best match and 0 is the worst match. "
                                  "The answer match score is based on the similarity of the two answers to a question.")
        
        # Stores the user prompt for the answer matching task
        USER_PROMPT = """You will receive questions with answer pairs to score for semantic similarity.

                        CRITICAL: Return ONLY valid JSON with this exact structure. Ensure all strings are properly escaped:
                        {
                            "results": [
                                {
                                    "q": "question text",
                                    "s": 0.85
                                }
                            ]
                        }

                        IMPORTANT JSON RULES:
                        - Escape all quotes, backslashes, and special characters in question text
                        - Use proper JSON number format for scores (0.0 to 1.0)
                        - Ensure the JSON is complete and well-formed
                        - Do not truncate the response - include all items

                        Scoring:
                        1.0 same meaning; 0.9–0.99 tiny nuance; 0.7–0.89 same stance but weaker/hedged; 0.4–0.69 partial; 0.1–0.39 mostly different; 0.0 contradiction/different entity/empty.
                        Rules: Identity same=1.0 else 0.0; Yes/No yes↔yes=1.0, yes↔no=0.0, yes↔"not yet/working on it"≈0.25–0.40; Numbers equal=1.0, ±10%=0.9, ±25%=0.7, ±50%=0.45, >50%=0.1; scope/timeframe/polarity mismatch → low/0; multi-part needs both parts.

                        Score each question-answer pair based on semantic similarity of a1 vs a2 given the question context.
                        """
        
        # Prepare the items payload for the AI request
        items_payload = {
            "questions": items_json
        }
        
        # Makes an AI request with the items for answer matching
        response = self._make_ai_request(
            user_prompt=USER_PROMPT,
            model=ANSWER_MATCHING_MODEL,
            system_prompt=ANSWER_MATCHING_PROMPT,
            temperature=ANSWER_MATCHING_TEMPERATURE,
            max_tokens=16000,
            has_arr_content=True,
            arr_content=items_payload
        )

        # Gets the response content and parse it
        resp_content = response.choices[0].message.content
        
        # Parse response content using OpenAI's native JSON parsing -- This is not decomposed given the nature of the task
        try:
            # Parse the JSON response directly (OpenAI's JSON mode ensures valid JSON)
            answer_scores_response = json.loads(resp_content)
            results = answer_scores_response.get('results', [])
            print(f"Successfully parsed answer match scores for {len(results)} items")
            
            # Extract scores and update the unanswered questionnaire
            answer_scores = {}
            for result in results:
                question = result.get('q')
                score = result.get('s', 0)
                answer_scores[question] = score
                
                # Set the answer match score in the unanswered questionnaire
                if question in self.unanswered_questionnaire.questions:
                    self.unanswered_questionnaire.questions[question].set_answer_match_score(score)
            
            # Returns the answer scores
            return answer_scores
        
        # Enhanced fallback parsing for malformed JSON
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response from AI model for answer matching: {e}")
            print(f"Full response content length: {len(resp_content)}")
            print(f"Response preview (first 500 chars): {resp_content[:500]}")
            print(f"Response ending (last 200 chars): {resp_content[-200:]}")
            return {}
                
        except Exception as e:
            print(f"Unexpected error during answer matching response processing: {e}")
            print(f"Error type: {type(e).__name__}")
            return {}
    

    
    """Fills the best matches from the reference questionnaire.
        
        Returns:
            best_matches (list): The best matches.
    """
    def _fill_best_matches(self):

        # Gets the best matches from the reference questionnaire
        best_matches = self._match_questions_to_reference()

        # Fills the best matches
        for match in best_matches:

            # Stores the unmatched question
            unmatched_question = match["unmatched_question"]

            # Stores the best match
            if (match["match"] != None):
                best_match = match["match"]["matched_question"]
            else:
                best_match = None

            # Sets the question match score
            if (match["match"] != None):
                self.unanswered_questionnaire.questions[unmatched_question].set_question_match_score(match["match"]["similarity"])
            else:
                self.unanswered_questionnaire.questions[unmatched_question].set_question_match_score(0)

            # Sets the reference question to be the best matching question from the reference questionnaire
            self.unanswered_questionnaire.questions[unmatched_question].set_reference_question(best_match)

        # Fill the answer match scores for all questions that have reference matches
        self._fill_answer_matches_score()

        # Returns the best matches
        return best_matches
    


    """Generates a combined questionnaire with conditional formatting for low match scores.
        
        Args:
            output_file_name (str): Output file name. Should end with .xlsx for Excel format with styling.
            If not specified, defaults to "combined_questionnaire.xlsx". If the specified file name ends with .csv,
            it will be saved as a CSV file without styling.
        
        Returns:
            None: Saves the combined questionnaire to the specified file.
    """
    def generate_combined_questionnaire(self, output_file_name="combined_questionnaire.xlsx"):

        # Creates a list to store the rows for the combined questionnaire
        combined_questionnaire_rows = []

        # Gets the questions from the unanswered questionnaire
        unanswered_questions = self.unanswered_questionnaire.get_questions()

        # Iterates through the unanswered questionnaire
        for question in unanswered_questions:

            # Gets the current question and answer
            current_question = unanswered_questions[question].get_question()
            current_answer = unanswered_questions[question].get_answer()

            # Gets the last question and answer
            last_question = unanswered_questions[question].get_reference_question()
            
            # Handle case where no match was found (last_question is None)
            if (last_question != None and last_question in self.reference_questionnaire.questions):
                question_id = self.reference_questionnaire.questions[last_question].get_question_id() + 1
                last_answer = self.reference_questionnaire.questions[last_question].get_answer()
            else:
                question_id = -1  # No match found, use -1 for error
                last_answer = ""

            # Gets the question match score
            question_match_score = unanswered_questions[question].get_question_match_score()

            # Gets the answer match score
            answer_match_score = unanswered_questions[question].get_answer_match_score()

            # Adds the question and answer data to the rows list
            combined_questionnaire_rows.append({
                "Current Question": current_question, 
                "Matched Question": last_question, 
                "Matched Question Row": question_id,
                "Question Match Score": question_match_score,
                "Current Answer": current_answer, 
                "Matched Answer": last_answer, 
                "Answer Match Score": answer_match_score
            })
            
        # Creates the combined questionnaire DataFrame from the collected rows
        combined_questionnaire = pd.DataFrame(combined_questionnaire_rows)
        
        # Check if output should be Excel format (for styling) or CSV
        if output_file_name.lower().endswith('.xlsx'):
            # Save to Excel with conditional formatting
            with pd.ExcelWriter(output_file_name, engine='openpyxl') as writer:
                combined_questionnaire.to_excel(writer, sheet_name='Combined Questionnaire', index=False)
                
                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Combined Questionnaire']
                
                # Define pink fill for cells with scores < self.accuracy_threshold
                pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
                
                # Define text wrapping alignment
                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths and formatting
                # Column mapping: A=Current Question, B=Matched Question, C=Matched Question Row, 
                # D=Question Match Score, E=Current Answer, F=Matched Answer, G=Answer Match Score
                text_columns = ['A', 'B', 'E', 'F']  # Text columns that need wider width and wrapping
                number_columns = ['C', 'D', 'G']     # Number columns (question ID and scores)
                
                # Set width and wrapping for text columns
                for col in text_columns:
                    worksheet.column_dimensions[col].width = 30
                    # Apply text wrapping to all cells in text columns
                    for row_idx in range(1, len(combined_questionnaire) + 2):  # Include header row
                        cell = worksheet.cell(row=row_idx, column=ord(col) - ord('A') + 1)
                        cell.alignment = wrap_alignment
                
                # Set appropriate width for number columns and apply wrapping
                for col in number_columns:
                    worksheet.column_dimensions[col].width = 12
                    # Apply text wrapping to all cells in number columns
                    for row_idx in range(1, len(combined_questionnaire) + 2):  # Include header row
                        cell = worksheet.cell(row=row_idx, column=ord(col) - ord('A') + 1)
                        cell.alignment = wrap_alignment
                
                # Apply conditional formatting to Question Match Score column (column D, index 4)
                question_score_col = 4  # 1-indexed (D column)
                for row_idx in range(2, len(combined_questionnaire) + 2):  # Start from row 2 (after header)
                    cell = worksheet.cell(row=row_idx, column=question_score_col)
                    if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < self.accuracy_threshold:
                        cell.fill = pink_fill
                
                # Apply conditional formatting to Answer Match Score column (column G, index 7)
                answer_score_col = 7  # 1-indexed (G column)
                for row_idx in range(2, len(combined_questionnaire) + 2):  # Start from row 2 (after header)
                    cell = worksheet.cell(row=row_idx, column=answer_score_col)
                    if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < self.accuracy_threshold:
                        cell.fill = pink_fill
        else:
            # Save as CSV (no styling possible)
            combined_questionnaire.to_csv(output_file_name, index=False)
            
        # Alerts the user that the combined questionnaire has been saved
        print(f"Combined questionnaire has been saved to {output_file_name}!")